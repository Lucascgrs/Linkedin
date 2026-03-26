"""
ExportUtils — utilitaires de persistance pour les données scrapées et les listes de suivi.

Fonctionnalités :
  - to_json / to_excel / to_json_and_excel : export des résultats de scraping
  - load_workbook  : charge un Excel existant → list[dict]
  - save_workbook  : écrit un Excel formaté (en-têtes colorés, colonnes fixes)
  - upsert         : insère ou met à jour un enregistrement par clé dans une liste
"""
import json
import logging
import os

logger = logging.getLogger(__name__)


class ExportUtils:
    """
    Utilitaires statiques pour exporter et persister des données.

    Les méthodes ``load_workbook``, ``save_workbook`` et ``upsert`` sont
    partagées par ``ConnectionManager`` et ``EasyApply`` afin d'éviter toute
    duplication de code.
    """

    @staticmethod
    def _ensure_dir(filepath: str) -> None:
        """Crée les dossiers parents si nécessaire."""
        directory = os.path.dirname(filepath)
        if directory:
            os.makedirs(directory, exist_ok=True)

    @staticmethod
    def to_json(data: list[dict], filepath: str) -> None:
        """
        Save a list of dicts to a JSON file.

        Args:
            data:     List of dicts to export.
            filepath: Destination file path (created if missing).
        """
        ExportUtils._ensure_dir(filepath)
        with open(filepath, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        print(f"💾 JSON sauvegardé : {filepath} ({len(data)} entrées)")

    @staticmethod
    def to_excel(
        data: list[dict],
        filepath: str,
        sheet_name: str = "Data",
    ) -> None:
        """
        Save a list of dicts to an Excel (.xlsx) file.

        Features:
        - Auto-adjusted column widths
        - First row frozen and bold

        Args:
            data:       List of dicts to export.
            filepath:   Destination file path (created if missing).
            sheet_name: Name of the worksheet.
        """
        try:
            import openpyxl
            from openpyxl.styles import Font
        except ImportError as exc:
            raise ImportError(
                "openpyxl is required for Excel export. "
                "Install it with: pip install openpyxl"
            ) from exc

        ExportUtils._ensure_dir(filepath)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name

        if not data:
            wb.save(filepath)
            print(f"💾 Excel sauvegardé : {filepath} (vide)")
            return

        headers = list(data[0].keys())
        ws.append(headers)

        # Bold + freeze header row
        for cell in ws[1]:
            cell.font = Font(bold=True)
        ws.freeze_panes = "A2"

        for row in data:
            cells = []
            for h in headers:
                val = row.get(h)
                if isinstance(val, (list, dict)):
                    val = json.dumps(val, ensure_ascii=False)
                cells.append(val)
            ws.append(cells)

        # Auto-adjust column widths
        for col in ws.columns:
            max_len = max(
                (len(str(cell.value)) if cell.value is not None else 0)
                for cell in col
            )
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 80)

        wb.save(filepath)
        print(f"💾 Excel sauvegardé : {filepath} ({len(data)} entrées)")

    @staticmethod
    def to_json_and_excel(
        data: list[dict],
        base_name: str,
        sheet_name: str = "Data",
    ) -> None:
        """
        Sauvegarde les données en JSON et Excel simultanément.

        Args:
            data:       Liste de dicts à exporter.
            base_name:  Chemin de base sans extension (ex. "output/companies").
            sheet_name: Nom de l'onglet Excel.
        """
        ExportUtils.to_json(data, f"{base_name}.json")
        ExportUtils.to_excel(data, f"{base_name}.xlsx", sheet_name)

    # ──────────────────────────────────────────────────────────────
    # Helpers partagés pour les fichiers de suivi (connexions, candidatures)
    # ──────────────────────────────────────────────────────────────

    @staticmethod
    def load_workbook(filepath: str) -> list[dict]:
        """
        Charge un fichier Excel de suivi et retourne une liste de dicts.

        Args:
            filepath: Chemin vers le fichier .xlsx.

        Returns:
            Liste de dicts (une entrée par ligne de données).
            Retourne [] si le fichier n'existe pas.
        """
        if not os.path.exists(filepath):
            return []
        try:
            import openpyxl
            wb = openpyxl.load_workbook(filepath)
            ws = wb.active
            headers = [cell.value for cell in ws[1]]
            rows = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if any(v is not None for v in row):
                    rows.append(dict(zip(headers, row)))
            return rows
        except Exception as exc:
            print(f"  ⚠️  Impossible de lire {filepath} : {exc}")
            logger.debug("load_workbook failed for %s", filepath, exc_info=True)
            return []

    @staticmethod
    def save_workbook(
        data: list[dict],
        filepath: str,
        sheet_name: str,
        columns: list[str],
        header_color: str = "1F4E79",
    ) -> None:
        """
        Écrit une liste de dicts dans un fichier Excel formaté.

        Crée les dossiers parents si nécessaire. Écrase le fichier existant
        (les données complètes sont toujours ré-écrites depuis la mémoire).

        Args:
            data:         Liste de dicts à sauvegarder.
            filepath:     Chemin de destination (.xlsx).
            sheet_name:   Nom de l'onglet.
            columns:      Ordre des colonnes (les clés manquantes donnent "").
            header_color: Couleur de fond des en-têtes (hex sans #).
        """
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill

        os.makedirs(os.path.dirname(filepath) or ".", exist_ok=True)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name

        # ── En-têtes ────────────────────────────────────────────
        ws.append(columns)
        header_fill = PatternFill("solid", fgColor=header_color)
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        ws.freeze_panes = "A2"

        # ── Données ─────────────────────────────────────────────
        for row in data:
            ws.append([row.get(col, "") for col in columns])

        # ── Largeurs automatiques ────────────────────────────────
        for col in ws.columns:
            max_len = max(
                (len(str(cell.value)) if cell.value is not None else 0)
                for cell in col
            )
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 70)

        wb.save(filepath)
        print(f"  💾 Excel mis à jour : {filepath} ({len(data)} entrée(s))")

    @staticmethod
    def upsert(records: list[dict], new_record: dict, key_field: str) -> list[dict]:
        """
        Insère ou met à jour un enregistrement dans une liste par clé.

        Args:
            records:    Liste existante de dicts.
            new_record: Enregistrement à insérer ou mettre à jour.
            key_field:  Nom du champ servant de clé primaire
                        (ex. "profile_url" ou "job_url").

        Returns:
            La liste mise à jour (modifiée en place + retournée).
        """
        key_value = new_record.get(key_field, "").rstrip("/")
        for i, record in enumerate(records):
            if record.get(key_field, "").rstrip("/") == key_value:
                records[i] = new_record
                return records
        records.append(new_record)
        return records
