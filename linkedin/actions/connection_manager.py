"""
ConnectionManager — envoie des demandes de connexion LinkedIn et tient un
fichier Excel persistant à jour avec les infos des personnes ajoutées.

Fonctionnalités :
  - add_connection(profile_url)          : ajoute une personne par URL
  - add_connections_bulk(urls)           : ajout en masse avec délai anti-ban
  - check_follow_back(profile_url)       : vérifie si la personne te suit
  - get_connections_df()                 : retourne le DataFrame courant

Fichier Excel de suivi :
  output/connections.xlsx  (créé si absent, mis à jour sinon — jamais remis à zéro)

Colonnes :
  profile_url | name | title | location | date_added | is_following_back | last_updated
"""
import asyncio
import os
import random
from datetime import datetime

# ---------------------------------------------------------------------------
# Helpers Excel (openpyxl — déjà utilisé dans le projet)
# ---------------------------------------------------------------------------
CONNECTIONS_FILE = os.path.join("output", "connections.xlsx")
SHEET_NAME = "Connections"
COLUMNS = [
    "profile_url",
    "name",
    "title",
    "location",
    "date_added",
    "is_following_back",
    "last_updated",
]


def _load_existing(filepath: str) -> list[dict]:
    """Charge le fichier Excel existant et retourne une liste de dicts.
    Retourne une liste vide si le fichier n'existe pas encore."""
    if not os.path.exists(filepath):
        return []
    try:
        import openpyxl
        wb = openpyxl.load_workbook(filepath)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if any(v is not None for v in row):
                rows.append(dict(zip(headers, row)))
        return rows
    except Exception as e:
        print(f"  ⚠️  Impossible de lire {filepath} : {e}")
        return []


def _save_excel(data: list[dict], filepath: str) -> None:
    """Sauvegarde la liste de dicts dans un fichier Excel.
    Crée le dossier parent si nécessaire."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    os.makedirs(os.path.dirname(filepath) or ".", exist_ok=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = SHEET_NAME

    # En-têtes
    ws.append(COLUMNS)
    header_fill = PatternFill("solid", fgColor="1F4E79")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"

    # Données
    for row in data:
        ws.append([row.get(col, "") for col in COLUMNS])

    # Largeurs auto
    for col in ws.columns:
        max_len = max(
            (len(str(cell.value)) if cell.value is not None else 0)
            for cell in col
        )
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)

    wb.save(filepath)
    print(f"  💾 Excel mis à jour : {filepath} ({len(data)} entrée(s))")


def _upsert(records: list[dict], new_record: dict) -> list[dict]:
    """Insère ou met à jour un enregistrement dans la liste selon profile_url."""
    url = new_record.get("profile_url", "").rstrip("/")
    for i, r in enumerate(records):
        if r.get("profile_url", "").rstrip("/") == url:
            records[i] = new_record  # mise à jour
            return records
    records.append(new_record)  # insertion
    return records


# ---------------------------------------------------------------------------
# ConnectionManager
# ---------------------------------------------------------------------------

class ConnectionManager:
    """
    Gère les demandes de connexion LinkedIn et tient un fichier Excel à jour.

    Usage :
        async with StealthBrowser(account="main") as browser:
            await SessionManager.load(browser)
            cm = ConnectionManager(browser.page)
            await cm.add_connection("https://www.linkedin.com/in/some-profile/")
    """

    def __init__(self, page, connections_file: str = CONNECTIONS_FILE) -> None:
        self.page = page
        self.connections_file = connections_file
        self._records: list[dict] = _load_existing(connections_file)
        print(f"  📂 {len(self._records)} contact(s) existant(s) chargé(s) depuis {connections_file}")

    # ------------------------------------------------------------------
    # Méthodes publiques
    # ------------------------------------------------------------------

    async def add_connection(
        self,
        profile_url: str,
        note: str = "",
    ) -> dict:
        """
        Navigue vers le profil LinkedIn et envoie une demande de connexion.

        Args:
            profile_url : URL complète du profil LinkedIn.
            note        : Message personnalisé optionnel joint à l'invitation
                          (max 300 caractères). Laisse vide pour envoyer sans note.

        Returns:
            Dict avec les infos scrapées + statut de l'action :
            {
                "profile_url", "name", "title", "location",
                "date_added", "is_following_back", "last_updated",
                "action_status"  # "sent" | "already_connected" | "already_sent" | "failed"
            }
        """
        profile_url = profile_url.rstrip("/") + "/"
        print(f"\n→ Traitement du profil : {profile_url}")

        try:
            await self.page.goto(profile_url, wait_until="domcontentloaded", timeout=30_000)
            await self._wait_for_profile()

            # Scrape des infos de base
            info = await self._scrape_profile_info()
            info["profile_url"] = profile_url

            # Vérification du suivi retour
            info["is_following_back"] = await self._check_following_back()

            # Tentative de connexion
            status = await self._send_connection_request(note=note)
            info["action_status"] = status

            # Horodatage
            now_str = datetime.now().strftime("%Y-%m-%d %H:%M")
            existing = self._find_existing(profile_url)
            info["date_added"] = existing.get("date_added", now_str) if existing else now_str
            info["last_updated"] = now_str

            # Persistance
            self._records = _upsert(self._records, {k: info.get(k, "") for k in COLUMNS})
            _save_excel(self._records, self.connections_file)

            _status_icon = {"sent": "✅", "already_connected": "🔗", "already_sent": "📨", "failed": "❌"}.get(status, "?")
            print(f"  {_status_icon} {info.get('name', '(inconnu)')} — {status}")
            return info

        except Exception as e:
            print(f"  ❌ Erreur sur {profile_url} : {e}")
            return {"profile_url": profile_url, "action_status": "failed", "error": str(e)}

    async def add_connections_bulk(
        self,
        profile_urls: list[str],
        note: str = "",
        delay_between: tuple = (8, 20),
        max_invitations: int = 20,
    ) -> list[dict]:
        """
        Envoie des demandes de connexion à une liste de profils.

        Args:
            profile_urls   : Liste d'URLs de profils LinkedIn.
            note           : Message optionnel joint à chaque invitation.
            delay_between  : (min_sec, max_sec) de pause entre chaque invitation.
            max_invitations: Nombre maximum d'invitations par session (sécurité anti-ban).

        Returns:
            Liste de dicts résultat pour chaque profil.
        """
        results = []
        targets = profile_urls[:max_invitations]

        for i, url in enumerate(targets, 1):
            print(f"\n  [{i:2d}/{len(targets)}] → {url}")
            result = await self.add_connection(url, note=note)
            results.append(result)

            if i < len(targets):
                delay = random.uniform(*delay_between)
                print(f"  ⏳ Pause de {delay:.1f}s...")
                await asyncio.sleep(delay)

        sent = sum(1 for r in results if r.get("action_status") == "sent")
        print(f"\n  ✓ Terminé : {sent} invitation(s) envoyée(s) sur {len(targets)} profil(s).")
        return results

    async def check_follow_back(self, profile_url: str) -> bool | None:
        """
        Vérifie si une personne te suit en naviguant sur son profil.
        Met à jour le fichier Excel si la personne est déjà dans la base.

        Returns:
            True / False si détecté, None si indéterminé.
        """
        profile_url = profile_url.rstrip("/") + "/"
        await self.page.goto(profile_url, wait_until="domcontentloaded", timeout=30_000)
        await self._wait_for_profile()
        follows = await self._check_following_back()

        existing = self._find_existing(profile_url)
        if existing:
            existing["is_following_back"] = follows
            existing["last_updated"] = datetime.now().strftime("%Y-%m-%d %H:%M")
            self._records = _upsert(self._records, existing)
            _save_excel(self._records, self.connections_file)

        return follows

    def get_connections_list(self) -> list[dict]:
        """Retourne la liste des contacts enregistrés."""
        return list(self._records)

    # ------------------------------------------------------------------
    # Méthodes privées — scraping & actions
    # ------------------------------------------------------------------

    async def _wait_for_profile(self) -> None:
        """Attend que la page profil soit chargée."""
        try:
            await self.page.wait_for_selector(
                "main, .scaffold-layout__main, section.artdeco-card",
                timeout=15_000,
            )
        except Exception:
            pass
        await asyncio.sleep(random.uniform(1.5, 2.5))

    async def _scrape_profile_info(self) -> dict:
        """Extrait nom, titre et localisation depuis la page profil."""
        info: dict = {"name": "", "title": "", "location": ""}

        # Nom
        for sel in [
            "h1.text-heading-xlarge",
            "h1[class*='inline']",
            ".pv-text-details__left-panel h1",
            "h1",
        ]:
            try:
                el = self.page.locator(sel).first
                if await el.count() > 0:
                    text = (await el.inner_text()).strip()
                    if text:
                        info["name"] = text
                        break
            except Exception:
                pass

        # Titre (headline)
        for sel in [
            ".text-body-medium.break-words",
            ".pv-text-details__left-panel .text-body-medium",
            "[data-generated-suggestion-target] .text-body-medium",
            ".ph5 .text-body-medium",
        ]:
            try:
                el = self.page.locator(sel).first
                if await el.count() > 0:
                    text = (await el.inner_text()).strip()
                    if text and len(text) > 3:
                        info["title"] = text
                        break
            except Exception:
                pass

        # Localisation
        for sel in [
            ".pv-text-details__left-panel span.text-body-small",
            ".pb2.pv-text-details__left-panel span.text-body-small",
            "span.text-body-small[aria-hidden='true']",
        ]:
            try:
                els = self.page.locator(sel)
                count = await els.count()
                for i in range(count):
                    text = (await els.nth(i).inner_text()).strip()
                    # La localisation ressemble à "Paris, Île-de-France, France"
                    if text and "," in text and len(text) < 80:
                        info["location"] = text
                        break
                if info["location"]:
                    break
            except Exception:
                pass

        print(f"  📋 Profil : {info['name']} | {info['title']} | {info['location']}")
        return info

    async def _check_following_back(self) -> bool | None:
        """
        Détecte si la personne suit ton profil.
        LinkedIn affiche parfois "Suit votre profil" ou "Follows you" sous le nom.
        Retourne True/False si détecté, None sinon (indéterminé).
        """
        following_indicators = [
            # Anglais
            "Follows you",
            "follows you",
            # Français
            "Suit votre profil",
            "suit votre profil",
            "Vous suit",
            "vous suit",
        ]
        try:
            page_text = await self.page.evaluate("() => document.body.innerText")
            for indicator in following_indicators:
                if indicator in page_text:
                    return True
            return False
        except Exception:
            return None

    async def _send_connection_request(self, note: str = "") -> str:
        """
        Clique sur le bouton "Se connecter" du profil via JS natif.
        Le clic JS bypass les vérifications de visibilité de Playwright et
        ne déplace pas la souris — fonctionne même avec le navigateur minimisé.

        Returns:
            "sent"              — invitation envoyée avec succès
            "already_connected" — déjà connecté (1er degré)
            "already_sent"      — invitation déjà en attente
            "failed"            — impossible d'envoyer
        """
        await self.page.evaluate("window.scrollTo(0, 0)")
        await asyncio.sleep(random.uniform(0.5, 1.0))

        # Le bouton du profil cible est toujours artdeco-button--primary.
        # Les boutons "Se connecter" dans les suggestions sont secondary → exclus.
        connect_selectors = [
            'button.artdeco-button--primary[aria-label*="rejoindre"]',
            'button.artdeco-button--primary[aria-label*="network" i]',
            'button.artdeco-button--primary:has(svg[data-test-icon="connect-small"])',
            '.pvs-sticky-header-profile-actions button[aria-label*="rejoindre"]',
            '.pvs-sticky-header-profile-actions button[aria-label*="network" i]',
        ]

        connect_found = False

        # Étape 1 : sélecteur CSS direct via JS .click()
        for sel in connect_selectors:
            try:
                if await self.page.locator(sel).count() > 0:
                    clicked = await self.page.evaluate(
                        "(sel) => { const el = document.querySelector(sel); if (el) { el.click(); return true; } return false; }",
                        sel,
                    )
                    if clicked:
                        connect_found = True
                        print(f"  [✓] Bouton 'Se connecter' cliqué.")
                        break
            except Exception:
                pass

        # Étape 2 : JS fallback — parcourt tous les boutons primaires
        if not connect_found:
            try:
                clicked = await self.page.evaluate("""
                    () => {
                        const btn = Array.from(document.querySelectorAll('button')).find(b => {
                            const lbl = (b.getAttribute('aria-label') || '').toLowerCase();
                            const cls = b.className || '';
                            return cls.includes('artdeco-button--primary') && (
                                lbl.includes('rejoindre') || lbl.includes('network') || lbl.includes('invitez')
                            );
                        });
                        if (btn) { btn.click(); return true; }
                        return false;
                    }
                """)
                if clicked:
                    connect_found = True
                    print("  [✓] Bouton 'Se connecter' cliqué (JS fallback).")
            except Exception:
                pass

        # Étape 3 : menu "Plus" (si le bouton est caché derrière)
        if not connect_found:
            more_selectors = [
                'button[aria-label*="Plus d" i]',
                'button[aria-label*="More" i]',
                '.pvs-sticky-header-profile-actions button:has-text("Plus")',
            ]
            clicked_more = False
            for sel in more_selectors:
                try:
                    if await self.page.locator(sel).count() > 0:
                        clicked = await self.page.evaluate(
                            "(sel) => { const el = document.querySelector(sel); if (el) { el.click(); return true; } return false; }",
                            sel,
                        )
                        if clicked:
                            clicked_more = True
                            await asyncio.sleep(random.uniform(0.8, 1.5))
                            break
                except Exception:
                    pass

            if clicked_more:
                dropdown_selectors = [
                    '.artdeco-dropdown__content li button[aria-label*="connecter" i]',
                    '.artdeco-dropdown__content li button[aria-label*="connect" i]',
                    '.artdeco-dropdown__content li button:has-text("Se connecter")',
                    '.artdeco-dropdown__content li button:has-text("Connect")',
                ]
                for sel in dropdown_selectors:
                    try:
                        if await self.page.locator(sel).count() > 0:
                            clicked = await self.page.evaluate(
                                "(sel) => { const el = document.querySelector(sel); if (el) { el.click(); return true; } return false; }",
                                sel,
                            )
                            if clicked:
                                connect_found = True
                                print("  [✓] Bouton 'Se connecter' cliqué via menu Plus.")
                                break
                    except Exception:
                        pass

        # Aucun bouton cliqué — déterminer le statut réel
        if not connect_found:
            try:
                page_text = await self.page.evaluate("() => document.body.innerText")
                if ("• 1er" in page_text or "· 1st" in page_text
                        or "1st degree" in page_text.lower()
                        or "1er degré" in page_text.lower()):
                    print("  [ℹ️] Déjà connecté (1er degré).")
                    return "already_connected"
            except Exception:
                pass

            pending_selectors = [
                'button[aria-label*="En attente" i]',
                'button[aria-label*="Pending" i]',
                'button:has-text("En attente")',
                'button:has-text("Pending")',
            ]
            for sel in pending_selectors:
                try:
                    if await self.page.locator(sel).count() > 0 and await self.page.locator(sel).is_visible():
                        print("  [ℹ️] Invitation déjà en attente.")
                        return "already_sent"
                except Exception:
                    pass

            print("  [✗] Bouton 'Se connecter' introuvable.")
            return "failed"

        await asyncio.sleep(random.uniform(1.0, 2.0))
        return await self._handle_connect_modal(note=note)

    async def _handle_connect_modal(self, note: str = "") -> str:
        """
        Gère la modale de confirmation qui apparaît après 'Se connecter'.
        Envoie sans note ou saisit la note si fournie.

        Returns:
            "sent" si l'invitation est envoyée, "failed" sinon.
        """
        await asyncio.sleep(random.uniform(0.8, 1.5))

        modal_present = await self.page.evaluate("""
            () => !!(
                document.querySelector('div[data-test-modal]') ||
                document.querySelector('div.artdeco-modal') ||
                document.querySelector('[role="dialog"]')
            )
        """)

        if not modal_present:
            return "sent"

        # Saisie de la note personnalisée
        if note:
            note_clicked = await self.page.evaluate("""
                () => {
                    const btn = Array.from(document.querySelectorAll('button')).find(b =>
                        (b.getAttribute('aria-label') || '').toLowerCase().includes('note') ||
                        b.innerText.includes('Ajouter une note') || b.innerText.includes('Add a note')
                    );
                    if (btn) { btn.click(); return true; }
                    return false;
                }
            """)
            if note_clicked:
                await asyncio.sleep(random.uniform(0.5, 1.0))
                for ts in ['textarea[name="message"]', 'textarea[id*="custom-message"]',
                           'div[data-test-modal] textarea', 'div.artdeco-modal textarea']:
                    try:
                        ta = self.page.locator(ts).first
                        if await ta.count() > 0:
                            await ta.focus()
                            await ta.type(note[:300], delay=random.randint(25, 55))
                            await asyncio.sleep(random.uniform(0.5, 1.0))
                            break
                    except Exception:
                        pass

        # Clic sur Envoyer via JS
        sent = await self.page.evaluate("""
            () => {
                for (const sel of [
                    'button[aria-label*="Envoyer maintenant" i]',
                    'button[aria-label*="Send now" i]',
                    'button[aria-label*="Envoyer sans note" i]',
                    'button[aria-label*="Send without a note" i]',
                ]) {
                    const el = document.querySelector(sel);
                    if (el) { el.click(); return true; }
                }
                const btn = Array.from(document.querySelectorAll(
                    'div[data-test-modal] button, div.artdeco-modal button, [role="dialog"] button'
                )).find(b => b.innerText.trim() === 'Envoyer' || b.innerText.trim() === 'Send');
                if (btn) { btn.click(); return true; }
                const sub = document.querySelector(
                    'div[data-test-modal] button[type="submit"], div.artdeco-modal button[type="submit"]'
                );
                if (sub) { sub.click(); return true; }
                return false;
            }
        """)

        if sent:
            await asyncio.sleep(random.uniform(1.0, 2.0))
            return "sent"

        # Fermeture propre si échec
        await self.page.evaluate("""
            () => {
                const btn = document.querySelector(
                    'button[aria-label*="Ignorer" i], button[aria-label*="Dismiss" i], ' +
                    'button[aria-label="Fermer"], button[aria-label="Close"]'
                );
                if (btn) btn.click();
            }
        """)
        print("  [✗] Impossible de valider l'envoi dans la modale.")
        return "failed"

    def _find_existing(self, profile_url: str) -> dict | None:
        """Cherche un enregistrement existant par profile_url."""
        url = profile_url.rstrip("/")
        for r in self._records:
            if r.get("profile_url", "").rstrip("/") == url:
                return r
        return None

